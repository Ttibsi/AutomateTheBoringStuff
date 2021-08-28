# Project 2 - Updating a spreadsheet

import openpyxl as xl

wb = xl.load_workbook('produceSales.xlsx')
sheet = wb['Sheet']

# Produce types and updated price
PRICE_UPDATES = {'Garlic': 3.07,
					'Celery':1.19,
					'Lemon': 1.27}

# Loop through rows and update prices
for rowNum in range(2, sheet.max_row):
	produceName = sheet.cell(row=rowNum, column=1).value

	if produceName in PRICE_UPDATES:
		sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')
